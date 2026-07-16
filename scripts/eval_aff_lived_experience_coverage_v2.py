#!/usr/bin/env python3
"""Evaluate multi-layer v2 concept coverage via dimension-dwelling lived-experience statements.

Pipeline:
  1. Load ``svarupa_concept_layer`` affinity per selected analytical layer (MySQL when
     configured, else static ``data/kg/*_concept_layer.v1.json``). Dimensions outside
     a layer's affinity are never sent to that layer's analyze API, and generate
     only builds statements for the union of selected-layer affinities.
  2. ``--step generate`` delegates to ``generate_dimension_dwelling_prompts`` /
     ``dimension_dwelling_prompt_template``: build per-concept dwelling prompts, run
     Bedrock for lived-experience statements, and write the Excel workbook.
  3. ``--step evaluate`` POSTs each statement to one or more layer analyze endpoints
     with ``options.force=true`` (only when the row's dimension is in that layer's
     ``svarupa_concept_layer`` affinity):
       - ``POST /v2/affect/analyze``
       - ``POST /v2/narrative-arc/analyze``
       - ``POST /v2/psycholinguistic/analyze``
       - ``POST /v2/metaphor/analyze``
  4. Writes a **new** multi-layer results workbook. For each selected layer, record
     returned attribute concepts and ``match`` when any equals the row's
     ``concept`` slug; otherwise ``no match`` (or ``not applicable`` when skipped).

Dependencies (beyond requirements.txt):
  pip install openpyxl pymysql

Usage:
  # Full pipeline (generate dwelling statements + evaluate all 4 layers)
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py

  # Generate lived-experience Excel only (via dimension-dwelling pipeline)
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py --step generate

  # Evaluate all layers against an existing dwelling workbook
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py --step evaluate

  # Evaluate a single layer
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py \\
      --step evaluate --layers affect

  # Generate + evaluate for one dimension only
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py \\
      --dimension trigunas

  # Evaluate an existing workbook, keeping only one dimension's rows
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py \\
      --step evaluate --dimension trigunas

  # Smoke: first 3 concepts
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage_v2.py --limit 3
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
from datetime import datetime
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _SCRIPTS_DIR.parent
sys.path.insert(0, str(_REPO_ROOT / "src"))
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from generate_dimension_dwelling_prompts import (  # noqa: E402
    LIVED_EXPERIENCE_COLUMNS,
    run_dimension_dwelling_generation,
)
from dimension_dwelling_prompt_template import LIVED_EXPERIENCE_MAX_TOKENS  # noqa: E402

logger = logging.getLogger("eval_aff_lived_experience_coverage_v2")

_EVAL_DIR = _REPO_ROOT / "data" / "eval"
DEFAULT_OUT = _EVAL_DIR / "dimension_dwelling_lived_experience.xlsx"
DEFAULT_EVAL_OUT = _EVAL_DIR / "aff_lived_experience_coverage_v2_layers.xlsx"
DEFAULT_PROMPTS_DIR = _EVAL_DIR / "dimension_dwelling_prompts"
EXCEL_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"


def with_run_timestamp(path: Path, *, stamp: str) -> Path:
    """Insert ``_<stamp>`` before the suffix: ``name.xlsx`` → ``name_<stamp>.xlsx``."""
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def new_run_timestamp() -> str:
    return datetime.now().strftime(EXCEL_TIMESTAMP_FORMAT)

# CLI / config keys → path segment under /v2/{segment}/analyze
LAYER_DEFS: dict[str, dict[str, str]] = {
    "affect": {
        "column_prefix": "affect",
        "path": "/v2/affect/analyze",
        "title": "Affect (AFF)",
        "layer_code": "AFF",
    },
    "narrative-arc": {
        "column_prefix": "narrative_arc",
        "path": "/v2/narrative-arc/analyze",
        "title": "Narrative-arc (NAR)",
        "layer_code": "NAR",
    },
    "psycholinguistic": {
        "column_prefix": "psycholinguistic",
        "path": "/v2/psycholinguistic/analyze",
        "title": "Psycholinguistic (PSY)",
        "layer_code": "PSY",
    },
    "metaphor": {
        "column_prefix": "metaphor",
        "path": "/v2/metaphor/analyze",
        "title": "Metaphor (MET)",
        "layer_code": "MET",
    },
}
ALL_LAYER_KEYS = tuple(LAYER_DEFS.keys())

# Seeded from sql/001_svarupa_dimensions_concepts.sql (fallback when MySQL is down).
_STATIC_DIMENSION_SLUGS: dict[int, str] = {
    1: "pancha_mahabhutas",
    2: "trigunas",
    3: "manasika_prakriti",
    4: "seven_layers_of_consciousness",
    5: "pancha_koshas",
    6: "subtle_energies",
    7: "phenomenology",
    8: "sthayibhavas",
    9: "vyabhicaribhavas",
    10: "paths_of_engagement",
    11: "ashtanga_yoga",
    12: "architecture_of_inner_life",
    13: "cyclical_evolution_of_consciousness",
    14: "yuga_cycles",
    15: "tridosha",
    16: "svabhava_svadharma",
    17: "three_lenses",
    18: "jyotish_shastra",
    19: "pancha_klesha",
    20: "pancha_vritti",
    21: "antarayas",
    22: "brahmaviharas",
    23: "sadhana_chatushtaya",
    24: "daivi_asuri_sampat",
    25: "upasanas_vidyas",
    26: "hatha_disciplines",
    27: "advaita_darshana",
    28: "ways_of_knowing",
    29: "bandha_moksha",
    30: "sadhana_practices",
    31: "reserved_d31",
}

NOT_APPLICABLE = "not applicable"

# Identity columns from the dimension-dwelling lived-experience workbook.
IDENTITY_HEADERS = list(LIVED_EXPERIENCE_COLUMNS)

# Per-layer result columns (suffixes after the layer column_prefix).
LAYER_RESULT_SUFFIXES = (
    "match_result",
    "matched_attributes",
    "api_attributes",
    "llm_primary_used",
    "abstained_dimensions",
    "api_error",
)

# Legacy single-layer evaluate columns (pre-multi-layer workbooks).
_LEGACY_AFFECT_HEADERS = (
    "match_result",
    "matched_attributes",
    "api_attributes",
    "llm_primary_used",
    "abstained_dimensions",
    "api_error",
)


def _layer_headers(layer_key: str) -> list[str]:
    prefix = LAYER_DEFS[layer_key]["column_prefix"]
    return [f"{prefix}_{suffix}" for suffix in LAYER_RESULT_SUFFIXES]


def result_headers(*, layers: tuple[str, ...] = ALL_LAYER_KEYS) -> list[str]:
    headers = list(IDENTITY_HEADERS)
    for layer_key in layers:
        headers.extend(_layer_headers(layer_key))
    return headers


HEADERS = result_headers()


def load_dimension_slug_maps() -> tuple[dict[str, int], dict[int, str]]:
    """Return ``(slug→id, id→slug)`` from MySQL when available, else SQL seed."""
    from svarupa_affect.infrastructure.config import Settings
    from svarupa_affect.infrastructure.kg.mysql_client import open_mysql

    settings = Settings.load()
    if settings.mysql_host and settings.mysql_database:
        try:
            conn = open_mysql(settings)
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT dimension_id, slug FROM svarupa_dimensions ORDER BY dimension_id"
                    )
                    rows = cur.fetchall()
            finally:
                conn.close()
            if rows:
                id_to_slug = {int(r["dimension_id"]): str(r["slug"]) for r in rows}
                slug_to_id = {slug: dim_id for dim_id, slug in id_to_slug.items()}
                logger.info(
                    "Loaded %d dimension slug(s) from MySQL (%s)",
                    len(slug_to_id),
                    settings.mysql_database,
                )
                return slug_to_id, id_to_slug
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Could not load svarupa_dimensions from MySQL (%s); using static slug map",
                exc,
            )

    id_to_slug = dict(_STATIC_DIMENSION_SLUGS)
    slug_to_id = {slug: dim_id for dim_id, slug in id_to_slug.items()}
    logger.info("Using static dimension slug map (%d entries)", len(slug_to_id))
    return slug_to_id, id_to_slug


def load_layer_concept_affinity(
    layers: tuple[str, ...],
) -> tuple[dict[str, frozenset[int]], dict[str, int]]:
    """Load ``svarupa_concept_layer`` affinity + concept→dimension_id for selected layers.

    Uses ``build_concept_registry`` (MySQL preferred, static JSON fallback).

    Returns ``(affinity_by_layer_key, concept_slug→dimension_id)``.
    """
    from svarupa_affect.infrastructure.kg.concept_registry import build_concept_registry

    affinity_by_layer: dict[str, frozenset[int]] = {}
    concept_to_dimension: dict[str, int] = {}
    for layer_key in layers:
        layer_code = LAYER_DEFS[layer_key]["layer_code"]
        registry = build_concept_registry(layer_code=layer_code)
        affinity = registry.affinity()
        affinity_by_layer[layer_key] = affinity
        for dim_id in affinity:
            for info in registry.concepts(dim_id):
                concept_to_dimension.setdefault(info.slug, dim_id)
        logger.info(
            "concept_layer %s (%s): affinity_dimensions=%s primary=%s concepts=%d",
            layer_key,
            layer_code,
            sorted(affinity),
            sorted(registry.primary_dimensions()),
            sum(len(registry.concepts(d)) for d in affinity),
        )
    return affinity_by_layer, concept_to_dimension


def applicable_dimension_slugs(
    *,
    layers: tuple[str, ...],
    affinity_by_layer: dict[str, frozenset[int]],
    id_to_slug: dict[int, str],
) -> frozenset[str]:
    """Union of dimension slugs appearing in selected layers' concept_layer affinity."""
    dim_ids: set[int] = set()
    for layer_key in layers:
        dim_ids |= set(affinity_by_layer.get(layer_key, frozenset()))
    slugs = frozenset(id_to_slug[dim_id] for dim_id in dim_ids if dim_id in id_to_slug)
    missing = sorted(dim_id for dim_id in dim_ids if dim_id not in id_to_slug)
    if missing:
        logger.warning(
            "concept_layer affinity has dimension_id(s) with no slug mapping: %s",
            missing,
        )
    return slugs


def resolve_dimension_id(
    *,
    dimension_slug: str,
    concept_slug: str,
    slug_to_id: dict[str, int],
    concept_to_dimension: dict[str, int],
) -> int | None:
    """Resolve a workbook row to a dimension_id via slug, then concept lookup."""
    needle = dimension_slug.strip().lower()
    if needle:
        if needle in slug_to_id:
            return slug_to_id[needle]
        # Tolerate singular/plural CLI typos (triguna → trigunas).
        if not needle.endswith("s") and f"{needle}s" in slug_to_id:
            return slug_to_id[f"{needle}s"]
        if needle.endswith("s") and needle[:-1] in slug_to_id:
            return slug_to_id[needle[:-1]]

    concept = concept_slug.strip()
    if concept and concept in concept_to_dimension:
        return concept_to_dimension[concept]
    return None


def layer_applicable_for_dimension(
    layer_key: str,
    dimension_id: int | None,
    affinity_by_layer: dict[str, frozenset[int]],
) -> bool:
    if dimension_id is None:
        return False
    return dimension_id in affinity_by_layer.get(layer_key, frozenset())


def _not_applicable_layer_values() -> list[object]:
    return [NOT_APPLICABLE, "", "", "", "", ""]


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel I/O. Install with: pip install openpyxl"
        ) from exc
    return openpyxl


def read_workbook(path: Path) -> tuple[list[str], list[list[object]]]:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value or "") for cell in ws[1]]
    rows: list[list[object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(cell is not None and str(cell).strip() for cell in row):
            rows.append(list(row))
    return headers, rows


def save_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    openpyxl = _require_openpyxl()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Layer coverage"
    ws.append(headers)
    for row in rows:
        ws.append(["" if cell is None else cell for cell in row])
    wb.save(path)


def _col_index(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _cell(row: list[object], idx: int | None) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _row_key_from_values(
    concept: object,
    status: object,
    statement_number: object,
    regional_perspective: object = "",
) -> tuple[str, str, str, str]:
    return (
        str(concept or ""),
        str(status or ""),
        str(statement_number or ""),
        str(regional_perspective or ""),
    )


def attributes_from_response(payload: dict) -> list[str]:
    """Collect attribute slugs from the analyze response."""
    attrs: list[str] = []
    for score in payload.get("attribute_scores", []):
        attr = score.get("attribute")
        if attr:
            attrs.append(str(attr))
    if attrs:
        return attrs
    for signal in payload.get("signals", []):
        for score in signal.get("attribute_scores", []):
            attr = score.get("attribute")
            if attr:
                attrs.append(str(attr))
    return attrs


def evaluate_match(expected_slug: str, response: dict) -> tuple[str, str]:
    from svarupa_affect.infrastructure.kg.concept_registry import canonical_slug

    expected = canonical_slug(expected_slug)
    found = attributes_from_response(response)
    matched = sorted({a for a in found if canonical_slug(a) == expected})
    if matched:
        return "match", ", ".join(matched)
    return "no match", ""


API_MAX_RETRIES = 3
API_RETRY_DELAY_S = 5.0


def call_v2_analyze_api(
    *,
    api_url: str,
    analysis_text: str,
    timeout_s: float,
    force: bool,
    max_retries: int = API_MAX_RETRIES,
    retry_delay_s: float = API_RETRY_DELAY_S,
) -> tuple[dict | None, str]:
    import httpx

    body = {
        "analysis_text": analysis_text,
        "options": {"force": force},
    }
    last_error = ""
    attempts = max(1, max_retries)

    for attempt in range(1, attempts + 1):
        try:
            with httpx.Client(timeout=timeout_s) as client:
                resp = client.post(
                    api_url, json=body, headers={"Content-Type": "application/json"}
                )
            if resp.status_code != 200:
                last_error = f"HTTP {resp.status_code}: {resp.text[:500]}"
                if attempt < attempts:
                    logger.warning(
                        "API non-200 (%s) attempt %d/%d for %s; retrying in %.0fs",
                        resp.status_code,
                        attempt,
                        attempts,
                        api_url,
                        retry_delay_s,
                    )
                    time.sleep(retry_delay_s)
                    continue
                return None, last_error

            payload = resp.json()
            if not isinstance(payload, dict):
                last_error = "API returned non-object JSON"
                if attempt < attempts:
                    logger.warning(
                        "API bad JSON attempt %d/%d for %s; retrying in %.0fs",
                        attempt,
                        attempts,
                        api_url,
                        retry_delay_s,
                    )
                    time.sleep(retry_delay_s)
                    continue
                return None, last_error
            return payload, ""
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
            if attempt < attempts:
                logger.warning(
                    "API error attempt %d/%d for %s: %s; retrying in %.0fs",
                    attempt,
                    attempts,
                    api_url,
                    last_error,
                    retry_delay_s,
                )
                time.sleep(retry_delay_s)
                continue
            return None, last_error

    return None, last_error or "API call failed after retries"


def layer_api_url(api_base: str, layer_key: str) -> str:
    base = api_base.rstrip("/")
    return f"{base}{LAYER_DEFS[layer_key]['path']}"


def _empty_layer_values() -> list[object]:
    return [""] * len(LAYER_RESULT_SUFFIXES)


def _apply_layer_response(
    *,
    concept_slug: str,
    response: dict | None,
    err: str,
) -> list[object]:
    """Return the six per-layer cell values."""
    if err:
        return ["", "", "", "", "", err]

    payload = response or {}
    result, matched = evaluate_match(concept_slug, payload)
    attrs = attributes_from_response(payload)
    abstained = payload.get("abstained_dimensions") or []
    return [
        result,
        matched,
        ", ".join(attrs),
        str(bool(payload.get("llm_primary_used"))),
        ", ".join(str(item) for item in abstained),
        "",
    ]


def _prior_layer_map(
    path: Path,
) -> dict[tuple[str, str, str, str], dict[str, list[object]]]:
    """Load existing multi-layer (or legacy) results keyed by identity."""
    if not path.is_file():
        return {}
    headers, rows = read_workbook(path)
    prior: dict[tuple[str, str, str, str], dict[str, list[object]]] = {}

    # Prefer dwelling-workbook keys; fall back to legacy AFF eval keys.
    concept_col = _col_index(headers, "concept")
    if concept_col is None:
        concept_col = _col_index(headers, "concept_slug")
    status_col = _col_index(headers, "status")
    if status_col is None:
        status_col = _col_index(headers, "state_code")
    stmt_col = _col_index(headers, "statement_number")
    if stmt_col is None:
        stmt_col = _col_index(headers, "question_number")
    perspective_col = _col_index(headers, "regional_perspective")

    if concept_col is None or stmt_col is None:
        return {}

    for row in rows:
        key = _row_key_from_values(
            row[concept_col],
            row[status_col] if status_col is not None else "",
            row[stmt_col],
            row[perspective_col] if perspective_col is not None else "",
        )
        layer_vals: dict[str, list[object]] = {}
        for layer_key in ALL_LAYER_KEYS:
            cols = _layer_headers(layer_key)
            if all(_col_index(headers, c) is not None for c in cols):
                layer_vals[layer_key] = [_cell(row, _col_index(headers, c)) for c in cols]
            elif layer_key == "affect" and all(
                _col_index(headers, c) is not None for c in _LEGACY_AFFECT_HEADERS
            ):
                layer_vals[layer_key] = [
                    _cell(row, _col_index(headers, c)) for c in _LEGACY_AFFECT_HEADERS
                ]
            else:
                layer_vals[layer_key] = _empty_layer_values()
        prior[key] = layer_vals
    return prior


def _limit_source_rows(
    headers: list[str],
    rows: list[list[object]],
    *,
    limit: int | None,
) -> list[list[object]]:
    """Keep rows for the first N distinct concepts when ``limit`` is set."""
    if limit is None:
        return rows
    concept_col = _col_index(headers, "concept")
    if concept_col is None:
        concept_col = _col_index(headers, "concept_slug")
    if concept_col is None:
        return rows[:limit]

    allowed: set[str] = set()
    out: list[list[object]] = []
    for row in rows:
        slug = _cell(row, concept_col)
        if slug not in allowed and len(allowed) >= limit:
            continue
        allowed.add(slug)
        out.append(row)
    return out


def _filter_source_rows(
    headers: list[str],
    rows: list[list[object]],
    *,
    dimension: str | None,
    limit: int | None,
) -> list[list[object]]:
    """Optionally keep one dimension, then the first N distinct concepts."""
    filtered = rows
    if dimension:
        dim_col = _col_index(headers, "dimension")
        if dim_col is None:
            dim_col = _col_index(headers, "dimension_name")
        if dim_col is None:
            raise RuntimeError(
                f"Cannot filter by dimension={dimension!r}: workbook has no dimension column."
            )
        wanted = dimension.strip().lower()
        filtered = [
            row
            for row in filtered
            if _cell(row, dim_col).lower() == wanted
        ]
        logger.info(
            "Dimension filter %r kept %d / %d row(s)",
            dimension,
            len(filtered),
            len(rows),
        )
        if not filtered:
            raise RuntimeError(
                f"No rows found for dimension={dimension!r}. "
                "Re-run --step generate with the same --dimension, or check the workbook."
            )

    return _limit_source_rows(headers, filtered, limit=limit)


def evaluate_workbook(
    source_path: Path,
    eval_path: Path,
    *,
    api_base: str,
    layers: tuple[str, ...],
    timeout_s: float,
    force: bool,
    api_force: bool,
    dimension: str | None = None,
    limit: int | None = None,
    affinity_by_layer: dict[str, frozenset[int]] | None = None,
    concept_to_dimension: dict[str, int] | None = None,
    slug_to_id: dict[str, int] | None = None,
) -> None:
    headers, source_rows = read_workbook(source_path)

    if affinity_by_layer is None or concept_to_dimension is None:
        affinity_by_layer, concept_to_dimension = load_layer_concept_affinity(layers)
    if slug_to_id is None:
        slug_to_id, _id_to_slug = load_dimension_slug_maps()

    # Accept dwelling workbook columns, or map legacy AFF-eval aliases.
    header_aliases = {
        "concept": ("concept", "concept_slug"),
        "statement": ("statement", "lived_experience_question"),
        "status": ("status", "state_code"),
        "statement_number": ("statement_number", "question_number"),
        "regional_perspective": ("regional_perspective",),
        "dimension": ("dimension", "dimension_name"),
        "display_name": ("display_name", "concept_name"),
    }

    def _resolve_header(logical: str) -> str | None:
        for candidate in header_aliases[logical]:
            if _col_index(headers, candidate) is not None:
                return candidate
        return None

    concept_header = _resolve_header("concept")
    statement_header = _resolve_header("statement")
    status_header = _resolve_header("status")
    number_header = _resolve_header("statement_number")
    perspective_header = _resolve_header("regional_perspective")
    dimension_header = _resolve_header("dimension")
    display_header = _resolve_header("display_name")

    if concept_header is None or statement_header is None or number_header is None:
        raise RuntimeError(
            "Workbook must contain concept + statement + statement_number columns "
            "(or legacy concept_slug + lived_experience_question + question_number)."
        )

    source_rows = _filter_source_rows(
        headers,
        source_rows,
        dimension=dimension,
        limit=limit,
    )

    prior = _prior_layer_map(eval_path)
    if source_path.resolve() != eval_path.resolve():
        for key, layer_vals in _prior_layer_map(source_path).items():
            prior.setdefault(key, layer_vals)

    out_headers = HEADERS
    out_rows: list[list[object]] = []
    updated = 0
    layer_stats = {
        lk: {
            "match": 0,
            "no match": 0,
            "error": 0,
            "skipped": 0,
            "not_applicable": 0,
        }
        for lk in layers
    }

    for source_row in source_rows:
        concept_slug = _cell(source_row, _col_index(headers, concept_header))
        question = _cell(source_row, _col_index(headers, statement_header))
        status = _cell(source_row, _col_index(headers, status_header)) if status_header else ""
        stmt_col = _col_index(headers, number_header)
        stmt_num = (
            source_row[stmt_col]
            if stmt_col is not None and stmt_col < len(source_row) and source_row[stmt_col] is not None
            else ""
        )
        perspective = (
            _cell(source_row, _col_index(headers, perspective_header))
            if perspective_header
            else ""
        )
        dimension_slug = (
            _cell(source_row, _col_index(headers, dimension_header)) if dimension_header else ""
        )
        identity = [
            dimension_slug,
            concept_slug,
            _cell(source_row, _col_index(headers, display_header)) if display_header else "",
            perspective,
            status,
            stmt_num,
            question,
        ]
        key = _row_key_from_values(concept_slug, status, stmt_num, perspective)
        prior_layers = prior.get(key, {lk: _empty_layer_values() for lk in ALL_LAYER_KEYS})
        layer_cells: dict[str, list[object]] = {
            lk: list(prior_layers.get(lk, _empty_layer_values())) for lk in ALL_LAYER_KEYS
        }

        dimension_id = resolve_dimension_id(
            dimension_slug=dimension_slug,
            concept_slug=concept_slug,
            slug_to_id=slug_to_id,
            concept_to_dimension=concept_to_dimension,
        )

        if not question:
            out_rows.append(identity + [v for lk in ALL_LAYER_KEYS for v in layer_cells[lk]])
            continue

        for layer_key in layers:
            already = str(layer_cells[layer_key][0] or "").strip()

            if not layer_applicable_for_dimension(layer_key, dimension_id, affinity_by_layer):
                if already != NOT_APPLICABLE:
                    layer_cells[layer_key] = _not_applicable_layer_values()
                    updated += 1
                layer_stats[layer_key]["not_applicable"] += 1
                logger.debug(
                    "Skip %s API: dimension=%r (id=%s) not in concept_layer affinity",
                    layer_key,
                    dimension_slug or concept_slug,
                    dimension_id,
                )
                continue

            if already and not force:
                layer_stats[layer_key]["skipped"] += 1
                continue

            response, err = call_v2_analyze_api(
                api_url=layer_api_url(api_base, layer_key),
                analysis_text=question,
                timeout_s=timeout_s,
                force=api_force,
            )
            layer_cells[layer_key] = _apply_layer_response(
                concept_slug=concept_slug,
                response=response,
                err=err,
            )
            updated += 1
            if err:
                layer_stats[layer_key]["error"] += 1
            else:
                result = str(layer_cells[layer_key][0])
                if result in layer_stats[layer_key]:
                    layer_stats[layer_key][result] += 1

        out_rows.append(identity + [v for lk in ALL_LAYER_KEYS for v in layer_cells[lk]])

        if updated and updated % 10 == 0:
            save_workbook(eval_path, out_headers, out_rows)
            logger.info("Checkpoint: %d layer-call(s) written to %s", updated, eval_path)

    save_workbook(eval_path, out_headers, out_rows)
    for layer_key in layers:
        stats = layer_stats[layer_key]
        logger.info(
            "%s: match=%d no_match=%d error=%d skipped=%d not_applicable=%d",
            LAYER_DEFS[layer_key]["title"],
            stats["match"],
            stats["no match"],
            stats["error"],
            stats["skipped"],
            stats["not_applicable"],
        )
    logger.info(
        "Evaluation complete: %d layer-call(s) processed → %s",
        updated,
        eval_path,
    )


def run_generate(
    output: Path,
    *,
    prompts_dir: Path,
    limit: int | None,
    dimension: str | None,
    concept: str | None,
    model_id: str | None,
    region: str | None,
    lived_experience_model_id: str | None,
    lived_experience_max_tokens: int,
    resume: bool,
    verbose: bool,
    allowed_dimension_slugs: frozenset[str] | None = None,
) -> Path:
    """Generate lived-experience Excel via the dimension-dwelling pipeline."""
    exit_code, excel_path = run_dimension_dwelling_generation(
        output_dir=prompts_dir,
        dimension=dimension,
        concept=concept,
        allowed_dimension_slugs=allowed_dimension_slugs,
        limit=limit,
        combined_only=False,
        generate_lived_experience=True,
        lived_experience_output=output,
        model_id=model_id,
        region=region,
        lived_experience_model_id=lived_experience_model_id,
        lived_experience_max_tokens=lived_experience_max_tokens,
        resume=resume,
        verbose=verbose,
        setup_logging=False,  # eval main() already configured logging
    )
    if exit_code != 0:
        raise SystemExit(f"Dimension-dwelling generation failed with exit code {exit_code}")
    if excel_path is None or not excel_path.is_file():
        raise SystemExit(f"Lived-experience workbook was not written: {output}")
    logger.info("Lived-experience workbook ready: %s", excel_path)
    return excel_path


def parse_layers(raw: list[str]) -> tuple[str, ...]:
    """Normalize ``--layers`` values; ``all`` expands to every layer."""
    if not raw:
        return ALL_LAYER_KEYS
    selected: list[str] = []
    for item in raw:
        token = item.strip().lower()
        if not token:
            continue
        if token == "all":
            return ALL_LAYER_KEYS
        aliases = {
            "aff": "affect",
            "nar": "narrative-arc",
            "narrative_arc": "narrative-arc",
            "narrative": "narrative-arc",
            "psy": "psycholinguistic",
            "met": "metaphor",
        }
        token = aliases.get(token, token)
        if token not in LAYER_DEFS:
            raise SystemExit(
                f"Unknown layer {item!r}. Choose from: all, {', '.join(ALL_LAYER_KEYS)}"
            )
        if token not in selected:
            selected.append(token)
    if not selected:
        return ALL_LAYER_KEYS
    return tuple(lk for lk in ALL_LAYER_KEYS if lk in selected)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--step",
        choices=("all", "generate", "evaluate"),
        default="all",
        help="run dwelling generation, API evaluation, or both (default: all)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUT,
        help=(
            "lived-experience questions workbook path "
            f"(default: {DEFAULT_OUT}; a _YYYYMMDD_HHMMSS stamp is appended on generate)"
        ),
    )
    parser.add_argument(
        "--eval-output",
        type=Path,
        default=DEFAULT_EVAL_OUT,
        help=(
            "multi-layer results workbook path "
            f"(default: {DEFAULT_EVAL_OUT}; a _YYYYMMDD_HHMMSS stamp is appended on evaluate)"
        ),
    )
    parser.add_argument(
        "--prompts-dir",
        type=Path,
        default=DEFAULT_PROMPTS_DIR,
        help=(
            "directory for per-concept dwelling prompt files "
            f"(default: {DEFAULT_PROMPTS_DIR})"
        ),
    )
    parser.add_argument(
        "--api-base",
        default="http://localhost:8000",
        help="API base URL (default: http://localhost:8000)",
    )
    parser.add_argument(
        "--api-url",
        default=None,
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--layers",
        nargs="+",
        default=["all"],
        metavar="LAYER",
        help=(
            "layers to evaluate: all, affect, narrative-arc, psycholinguistic, metaphor "
            "(aliases: aff, nar, psy, met). Default: all"
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="process only the first N concepts",
    )
    parser.add_argument(
        "--dimension",
        type=str,
        default="",
        help=(
            "dimension slug filter: generate only this dimension's lived-experience "
            "questions, and evaluate only rows for this dimension. Must be in "
            "svarupa_concept_layer affinity for at least one selected --layers value"
        ),
    )
    parser.add_argument(
        "--concept",
        type=str,
        default="",
        help="optional filter for generate: only this concept slug",
    )
    parser.add_argument(
        "--model-id",
        type=str,
        default="",
        help="Bedrock model for dwelling Step-2 / lived-experience generation",
    )
    parser.add_argument(
        "--region",
        type=str,
        default="",
        help="AWS region for Bedrock (default: AWS_REGION / us-west-2)",
    )
    parser.add_argument(
        "--lived-experience-model-id",
        type=str,
        default="",
        help="Bedrock model for lived-experience generation (default: --model-id)",
    )
    parser.add_argument(
        "--lived-experience-max-tokens",
        type=int,
        default=LIVED_EXPERIENCE_MAX_TOKENS,
        help=f"Max tokens for lived-experience generation (default: {LIVED_EXPERIENCE_MAX_TOKENS})",
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="ignore dwelling lived-experience checkpoint and regenerate every concept",
    )
    parser.add_argument(
        "--api-timeout",
        type=float,
        default=120.0,
        help="HTTP timeout per analyze call in seconds (default: 120)",
    )
    parser.add_argument(
        "--no-api-force",
        action="store_true",
        help="do not set options.force=true (salience gate may abstain)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="re-evaluate selected layers even when match_result is already set",
    )
    parser.add_argument("-v", "--verbose", action="store_true", help="debug logging")
    return parser.parse_args(argv)


def _resolve_api_base(args: argparse.Namespace) -> str:
    if args.api_url:
        url = str(args.api_url).rstrip("/")
        for suffix in ("/v2/affect/analyze", "/affect/analyze"):
            if url.endswith(suffix):
                return url[: -len(suffix)] or "http://localhost:8000"
        return url
    return str(args.api_base).rstrip("/")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )
    layers = parse_layers(args.layers)
    api_base = _resolve_api_base(args)
    dimension = args.dimension.strip() or None

    # Always load concept_layer applicability before generate / evaluate.
    slug_to_id, id_to_slug = load_dimension_slug_maps()
    affinity_by_layer, concept_to_dimension = load_layer_concept_affinity(layers)
    allowed_slugs = applicable_dimension_slugs(
        layers=layers,
        affinity_by_layer=affinity_by_layer,
        id_to_slug=id_to_slug,
    )
    logger.info(
        "Selected layers %s → concept_layer applicable dimensions: %s",
        ", ".join(layers),
        ", ".join(sorted(allowed_slugs)) or "(none)",
    )

    if dimension:
        dim_id = resolve_dimension_id(
            dimension_slug=dimension,
            concept_slug="",
            slug_to_id=slug_to_id,
            concept_to_dimension=concept_to_dimension,
        )
        resolved_slug = id_to_slug.get(dim_id) if dim_id is not None else None
        if dim_id is None or resolved_slug not in allowed_slugs:
            applicable_layers = [
                lk
                for lk in layers
                if dim_id is not None
                and layer_applicable_for_dimension(lk, dim_id, affinity_by_layer)
            ]
            raise SystemExit(
                f"Dimension {dimension!r} is not in svarupa_concept_layer affinity "
                f"for selected layer(s) {', '.join(layers)}. "
                f"Applicable dimensions: {', '.join(sorted(allowed_slugs)) or '(none)'}."
                + (
                    f" (resolved id={dim_id}, overlapping layers={applicable_layers})"
                    if dim_id is not None
                    else ""
                )
            )
        # Normalize CLI filter to the canonical DB slug (e.g. triguna → trigunas).
        dimension = resolved_slug

    # Stamp every Excel this run writes so successive executions do not overwrite.
    run_stamp = new_run_timestamp()
    do_generate = args.step in ("all", "generate")
    do_evaluate = args.step in ("all", "evaluate")
    questions_path = (
        with_run_timestamp(args.output, stamp=run_stamp) if do_generate else args.output
    )
    eval_path = (
        with_run_timestamp(args.eval_output, stamp=run_stamp) if do_evaluate else args.eval_output
    )
    if do_generate:
        logger.info("Lived-experience workbook will be written to %s", questions_path)
    if do_evaluate:
        logger.info("Results workbook will be written to %s", eval_path)

    if do_generate:
        if dimension:
            logger.info("Generating lived-experience questions for dimension=%r", dimension)
        else:
            logger.info(
                "Generating lived-experience questions for %d concept_layer dimension(s)",
                len(allowed_slugs),
            )
        run_generate(
            questions_path,
            prompts_dir=args.prompts_dir,
            limit=args.limit,
            dimension=dimension,
            concept=args.concept.strip() or None,
            model_id=args.model_id.strip() or None,
            region=args.region.strip() or None,
            lived_experience_model_id=args.lived_experience_model_id.strip() or None,
            lived_experience_max_tokens=args.lived_experience_max_tokens,
            resume=not args.no_resume,
            verbose=args.verbose,
            allowed_dimension_slugs=None if dimension else allowed_slugs,
        )

    if do_evaluate:
        if not questions_path.is_file():
            raise SystemExit(
                f"Workbook not found: {questions_path}. Run with --step generate first "
                "(or pass --output to an existing stamped workbook)."
            )
        logger.info(
            "Evaluating layers %s via %s → %s%s",
            ", ".join(layers),
            api_base,
            eval_path,
            f" (dimension={dimension!r})" if dimension else "",
        )
        evaluate_workbook(
            questions_path,
            eval_path,
            api_base=api_base,
            layers=layers,
            timeout_s=args.api_timeout,
            force=args.force,
            api_force=not args.no_api_force,
            dimension=dimension,
            # On --step all, generate already scoped the workbook; still pass
            # filters so evaluate-only and mixed workbooks stay correct.
            limit=args.limit,
            affinity_by_layer=affinity_by_layer,
            concept_to_dimension=concept_to_dimension,
            slug_to_id=slug_to_id,
        )

    print(f"Done. Questions: {questions_path.resolve()}")
    if do_evaluate:
        print(f"Results:  {eval_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
