#!/usr/bin/env python3
"""Evaluate AFF concept coverage via lived-experience questions.

Pipeline:
  1. Load **primary** dimension/concept pairs for layer AFF from ``svarupa_concept_layer``
     (``layer_code = 'AFF'`` and ``role = 'primary'``), or the same filter on
     ``data/kg/aff_concept_layer.v1.json`` when MySQL is not configured.
  2. Generate two first-person lived-experience questions per pair (Bedrock batch or
     template fallback) and write rows to an Excel workbook.
  3. POST each question to ``POST /analyze`` with ``{"analysis_text": "..."}``.
  4. Mark the row ``match`` when any ``attribute_scores[].attribute`` equals the row's
     ``concept_slug``; otherwise ``no match``.

Dependencies (beyond requirements.txt):
  pip install openpyxl

Usage:
  # Full pipeline (generate questions + call API + update Excel)
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage.py

  # Generate Excel only
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage.py --step generate

  # Evaluate an existing workbook (API must be running on :8000)
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage.py --step evaluate

  # Smoke test on first 5 concepts
  PYTHONPATH=src python scripts/eval_aff_lived_experience_coverage.py --limit 5
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from svarupa_affect import LAYER_CODE
from svarupa_affect.infrastructure.config import Settings

logger = logging.getLogger("eval_aff_lived_experience_coverage")

DEFAULT_OUT = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "eval"
    / "aff_lived_experience_coverage.xlsx"
)
STATIC_CONCEPT_LAYER = (
    Path(__file__).resolve().parents[1] / "data" / "kg" / "aff_concept_layer.v1.json"
)

HEADERS = [
    "dimension_id",
    "dimension_name",
    "concept_id",
    "concept_slug",
    "concept_name",
    "role",
    "question_number",
    "lived_experience_question",
    "match_result",
    "matched_attributes",
    "api_attributes",
    "api_error",
]

QUESTION_GEN_SYSTEM = (
    "You write first-person lived-experience questions for affect recognition research. "
    "Each question is a short natural English narrative (1–3 sentences) that a real person "
    "might say when their inner experience authentically expresses the given concept. "
    "Do not diagnose or label the person. Do not mention internal concept slugs. "
    "Output JSON only."
)

_OVERVIEW_RE = re.compile(r"Overview:\s*(.+?)(?:\n\n|\Z)", re.DOTALL)


@dataclass(frozen=True)
class ConceptRow:
    concept_id: int
    dimension_id: int
    slug: str
    name: str
    role: str
    gloss: str


def _gloss_excerpt(gloss: str, *, max_len: int = 320) -> str:
    match = _OVERVIEW_RE.search(gloss)
    text = match.group(1).strip() if match else gloss.strip().replace("\n", " ")
    if len(text) > max_len:
        return text[: max_len - 3].rstrip() + "..."
    return text


def _dimension_registry():
    from svarupa_affect.infrastructure.kg.dimension_registry import build_dimension_registry

    return build_dimension_registry()


def _dimension_name(dimension_id: int) -> str:
    try:
        return _dimension_registry().name_for(dimension_id)
    except Exception:
        return f"D{dimension_id}"


def load_aff_concepts(*, limit: int | None = None) -> list[ConceptRow]:
    """AFF primary dimension/concept pairs from MySQL or static JSON fallback."""
    settings = Settings.load()
    rows: list[ConceptRow] = []

    if settings.mysql_host and settings.mysql_database:
        try:
            from svarupa_affect.infrastructure.kg.mysql_client import open_mysql
        except ImportError:
            logger.warning("pymysql not installed; falling back to static concept layer JSON")
        else:
            sql = """
            SELECT c.concept_id, cl.dimension_id, c.slug, c.name, cl.role,
                   COALESCE(NULLIF(TRIM(c.description), ''), c.name) AS gloss
              FROM svarupa_concept_layer cl
              JOIN svarupa_concepts c ON c.concept_id = cl.concept_id
             WHERE cl.layer_code = %s
               AND cl.role = 'primary'
             ORDER BY cl.dimension_id, c.slug
        """
            conn = open_mysql(settings)
            try:
                with conn.cursor() as cur:
                    cur.execute(sql, (LAYER_CODE,))
                    for row in cur.fetchall():
                        rows.append(
                            ConceptRow(
                                concept_id=int(row["concept_id"]),
                                dimension_id=int(row["dimension_id"]),
                                slug=str(row["slug"]),
                                name=str(row["name"]),
                                role=str(row["role"]),
                                gloss=str(row["gloss"] or row["name"]),
                            )
                        )
            finally:
                conn.close()
            source = f"{settings.mysql_database}.svarupa_concept_layer (AFF, primary)"
            if limit is not None:
                rows = rows[:limit]
            logger.info("Loaded %d concept(s) from %s", len(rows), source)
            return rows

    if STATIC_CONCEPT_LAYER.is_file():
        payload = json.loads(STATIC_CONCEPT_LAYER.read_text(encoding="utf-8"))
        for item in payload.get("concepts", []):
            if str(item.get("role", "")).lower() != "primary":
                continue
            rows.append(
                ConceptRow(
                    concept_id=int(item["concept_id"]),
                    dimension_id=int(item["dimension_id"]),
                    slug=str(item["slug"]),
                    name=str(item["name"]),
                    role=str(item["role"]),
                    gloss=str(item.get("gloss") or item["name"]),
                )
            )
        source = f"{STATIC_CONCEPT_LAYER} (AFF, primary)"
    else:
        raise RuntimeError(
            "MySQL is not configured and static snapshot "
            f"{STATIC_CONCEPT_LAYER} is missing. Configure SVARUPA_MYSQL_* or export the snapshot."
        )

    if limit is not None:
        rows = rows[:limit]
    logger.info("Loaded %d concept(s) from %s", len(rows), source)
    return rows


def _template_questions(concept: ConceptRow, dim_name: str) -> tuple[str, str]:
    excerpt = _gloss_excerpt(concept.gloss)
    q1 = (
        f"Lately in my daily life I notice something like {concept.name.lower()}: "
        f"{excerpt} Can you help me understand what might be showing up in my experience?"
    )
    q2 = (
        f"When I sit with how I have been feeling, it reminds me of {concept.name.lower()} "
        f"within {dim_name}. {excerpt} This is what it has been like for me recently."
    )
    return q1, q2


async def _generate_questions_llm_batch(
    batch: list[ConceptRow],
    dim_names: dict[int, str],
    *,
    model_id: str,
    region: str | None,
    timeout_s: float,
    max_tokens: int,
) -> dict[int, tuple[str, str]]:
    from svarupa_affect.infrastructure.llm.bedrock_provider import BedrockLLMProvider

    if not region:
        raise RuntimeError("AWS region is not configured (set AWS_REGION or SVARUPA_AWS_REGION)")

    lines = []
    for concept in batch:
        dim_name = dim_names.get(concept.dimension_id, f"D{concept.dimension_id}")
        lines.append(
            f"- concept_id={concept.concept_id}; dimension_id={concept.dimension_id} "
            f"({dim_name}); slug={concept.slug}; name={concept.name}; role={concept.role}; "
            f"gloss_excerpt={_gloss_excerpt(concept.gloss)!r}"
        )
    prompt = (
        "For each concept below, write exactly two distinct lived_experience_question strings.\n"
        'Return JSON: {"items": [{"concept_id": <int>, "questions": ["q1", "q2"]}, ...]}\n\n'
        "Concepts:\n" + "\n".join(lines)
    )
    schema = {
        "type": "object",
        "properties": {
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "concept_id": {"type": "integer"},
                        "questions": {
                            "type": "array",
                            "items": {"type": "string"},
                            "minItems": 2,
                            "maxItems": 2,
                        },
                    },
                    "required": ["concept_id", "questions"],
                },
            }
        },
        "required": ["items"],
    }

    provider = BedrockLLMProvider(region_name=region, read_timeout_s=timeout_s + 5.0)
    payload = await provider.complete_json(
        system=QUESTION_GEN_SYSTEM,
        prompt=prompt,
        schema=schema,
        model_id=model_id,
        temperature=0.7,
        timeout_s=timeout_s,
        max_tokens=max_tokens,
    )

    out: dict[int, tuple[str, str]] = {}
    for item in payload.get("items", []):
        cid = int(item["concept_id"])
        questions = [str(q).strip() for q in item.get("questions", []) if str(q).strip()]
        if len(questions) >= 2:
            out[cid] = (questions[0], questions[1])
    return out


async def generate_all_questions(
    concepts: list[ConceptRow],
    *,
    use_templates: bool,
    batch_size: int,
) -> dict[int, tuple[str, str]]:
    dim_names = {c.dimension_id: _dimension_name(c.dimension_id) for c in concepts}

    if use_templates:
        return {c.concept_id: _template_questions(c, dim_names[c.dimension_id]) for c in concepts}

    settings = Settings.load()
    model_id = settings.bedrock_model_id
    region = settings.aws_region
    timeout_s = settings.llm_assist_timeout_s
    max_tokens = max(2048, settings.llm_assist_max_tokens)

    generated: dict[int, tuple[str, str]] = {}
    for start in range(0, len(concepts), batch_size):
        batch = concepts[start : start + batch_size]
        try:
            batch_out = await _generate_questions_llm_batch(
                batch,
                dim_names,
                model_id=model_id,
                region=region,
                timeout_s=timeout_s,
                max_tokens=max_tokens,
            )
        except Exception:
            logger.exception(
                "LLM batch failed for concepts %s–%s; using templates for that batch",
                start + 1,
                start + len(batch),
            )
            batch_out = {}

        for concept in batch:
            pair = batch_out.get(concept.concept_id)
            if pair is None:
                pair = _template_questions(concept, dim_names[concept.dimension_id])
            generated[concept.concept_id] = pair
        logger.info("Generated questions for %d / %d concepts", len(generated), len(concepts))

    return generated


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel I/O. Install with: pip install openpyxl"
        ) from exc
    return openpyxl


def build_workbook_rows(
    concepts: list[ConceptRow], questions: dict[int, tuple[str, str]]
) -> list[list[object]]:
    rows: list[list[object]] = []
    for concept in concepts:
        q1, q2 = questions[concept.concept_id]
        dim_name = _dimension_name(concept.dimension_id)
        for num, question in ((1, q1), (2, q2)):
            rows.append(
                [
                    concept.dimension_id,
                    dim_name,
                    concept.concept_id,
                    concept.slug,
                    concept.name,
                    concept.role,
                    num,
                    question,
                    "",
                    "",
                    "",
                    "",
                ]
            )
    return rows


def write_workbook(path: Path, data_rows: list[list[object]]) -> None:
    openpyxl = _require_openpyxl()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AFF coverage"
    ws.append(HEADERS)
    for row in data_rows:
        ws.append(row)
    wb.save(path)
    logger.info("Wrote %d row(s) to %s", len(data_rows), path)


def read_workbook(path: Path) -> tuple[list[str], list[list[object]]]:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value or "") for cell in ws[1]]
    rows: list[list[object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(cell is not None and str(cell).strip() for cell in row):
            rows.append(list(row))
    return headers, rows


def save_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AFF coverage"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _col_index(headers: list[str], name: str) -> int:
    try:
        return headers.index(name)
    except ValueError as exc:
        raise RuntimeError(f"Workbook missing required column {name!r}") from exc


def attributes_from_response(payload: dict, *, dimension_id: int | None = None) -> list[str]:
    """Collect attribute slugs; when ``dimension_id`` is set, use that signal only."""
    from svarupa_affect.infrastructure.kg.dimension_registry import StaticDimensionRegistry

    if dimension_id is not None:
        dim_name = StaticDimensionRegistry().name_for(dimension_id)
        for signal in payload.get("signals", []):
            if signal.get("dimension_name") == dim_name:
                return [
                    str(score.get("attribute"))
                    for score in signal.get("attribute_scores", [])
                    if score.get("attribute")
                ]
        return []

    attrs: list[str] = []
    for score in payload.get("attribute_scores", []):
        attr = score.get("attribute")
        if attr:
            attrs.append(str(attr))
    if attrs:
        return attrs
    for signal in payload.get("signals", []):
        for score in signal.get("attribute_scores", []):
            attr = score.get("attribute")
            if attr:
                attrs.append(str(attr))
    return attrs


def evaluate_match(
    expected_slug: str, response: dict, *, dimension_id: int | None = None
) -> tuple[str, str]:
    from svarupa_affect.infrastructure.kg.concept_registry import canonical_slug

    expected = canonical_slug(expected_slug)
    found = attributes_from_response(response, dimension_id=dimension_id)
    matched = sorted({a for a in found if canonical_slug(a) == expected})
    if matched:
        return "match", ", ".join(matched)
    return "no match", ""


def call_analyze_api(
    *,
    api_url: str,
    analysis_text: str,
    timeout_s: float,
) -> tuple[dict | None, str]:
    import httpx

    body = {"analysis_text": analysis_text}
    try:
        with httpx.Client(timeout=timeout_s) as client:
            resp = client.post(api_url, json=body, headers={"Content-Type": "application/json"})
        resp.raise_for_status()
        payload = resp.json()
        if not isinstance(payload, dict):
            return None, "API returned non-object JSON"
        return payload, ""
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)


def evaluate_workbook(
    path: Path,
    *,
    api_url: str,
    timeout_s: float,
    force: bool,
) -> None:
    headers, rows = read_workbook(path)
    idx = {name: _col_index(headers, name) for name in HEADERS if name in headers}
    for required in ("concept_slug", "lived_experience_question", "match_result"):
        if required not in idx:
            raise RuntimeError(f"Workbook missing required column {required!r}")

    updated = 0
    for row in rows:
        while len(row) < len(HEADERS):
            row.append("")
        question = str(row[idx["lived_experience_question"]] or "").strip()
        if not question:
            continue
        if not force and str(row[idx["match_result"]] or "").strip():
            continue

        concept_slug = str(row[idx["concept_slug"]] or "").strip()
        dimension_id = int(row[idx["dimension_id"]]) if row[idx["dimension_id"]] else None
        response, err = call_analyze_api(
            api_url=api_url, analysis_text=question, timeout_s=timeout_s
        )
        if err:
            row[idx["match_result"]] = ""
            row[_col_index(headers, "matched_attributes")] = ""
            row[_col_index(headers, "api_attributes")] = ""
            row[_col_index(headers, "api_error")] = err
        else:
            result, matched = evaluate_match(
                concept_slug, response or {}, dimension_id=dimension_id
            )
            row[idx["match_result"]] = result
            row[_col_index(headers, "matched_attributes")] = matched
            row[_col_index(headers, "api_attributes")] = ", ".join(
                attributes_from_response(response or {}, dimension_id=dimension_id)
            )
            row[_col_index(headers, "api_error")] = ""
        updated += 1
        if updated % 10 == 0:
            save_workbook(path, headers, rows)
            logger.info("Checkpoint: evaluated %d row(s)", updated)

    save_workbook(path, headers, rows)
    matches = sum(1 for row in rows if str(row[idx["match_result"]]) == "match")
    logger.info(
        "Evaluation complete: %d row(s) processed, %d match, %d no match",
        updated,
        matches,
        sum(1 for row in rows if str(row[idx["match_result"]]) == "no match"),
    )


async def run_generate(
    path: Path, *, limit: int | None, use_templates: bool, batch_size: int
) -> None:
    concepts = load_aff_concepts(limit=limit)
    questions = await generate_all_questions(
        concepts, use_templates=use_templates, batch_size=batch_size
    )
    rows = build_workbook_rows(concepts, questions)
    write_workbook(path, rows)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--step",
        choices=("all", "generate", "evaluate"),
        default="all",
        help="run question generation, API evaluation, or both (default: all)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUT,
        help=f"Excel workbook path (default: {DEFAULT_OUT})",
    )
    parser.add_argument(
        "--api-url",
        default="http://localhost:8000/analyze",
        help="AFF analyze endpoint (default: http://localhost:8000/analyze)",
    )
    parser.add_argument("--limit", type=int, default=None, help="process only the first N concepts")
    parser.add_argument(
        "--use-templates",
        action="store_true",
        help="skip Bedrock; build template questions from concept gloss",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=8,
        help="concepts per Bedrock batch when generating questions (default: 8)",
    )
    parser.add_argument(
        "--api-timeout",
        type=float,
        default=120.0,
        help="HTTP timeout per /analyze call in seconds (default: 120)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="re-evaluate rows even when match_result is already set",
    )
    parser.add_argument("-v", "--verbose", action="store_true", help="debug logging")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    if args.step in ("all", "generate"):
        asyncio.run(
            run_generate(
                args.output,
                limit=args.limit,
                use_templates=args.use_templates,
                batch_size=args.batch_size,
            )
        )

    if args.step in ("all", "evaluate"):
        if not args.output.is_file():
            raise SystemExit(f"Workbook not found: {args.output}. Run with --step generate first.")
        evaluate_workbook(
            args.output,
            api_url=args.api_url,
            timeout_s=args.api_timeout,
            force=args.force,
        )

    print(f"Done. Workbook: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
